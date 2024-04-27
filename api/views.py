import json
from datetime import datetime
from django.contrib.auth import authenticate, login, logout
from django.shortcuts import render
from jsonschema import ValidationError
from openpyxl import load_workbook
from django.contrib.auth import models as usr
from django.http import HttpResponse, JsonResponse
from django.core import serializers
from .models import *
from django.views.decorators.csrf import csrf_exempt,csrf_protect
from django.contrib import messages
from openpyxl import load_workbook
from .models import Department
from django.db import transaction
import pandas as pd
 
 

# Create your views here.
@csrf_exempt
def import_from_excel(request):
    if request.method == 'POST':
        try:
            excel_file = request.FILES['excel_file']
            # student_photo = request.FILES['photo']
            # father_photo = request.FILES['father_photo']
            # mother_photo = request.FILES['mother_photo']
            wb = load_workbook(excel_file)
            ws = wb.active
            students_to_create = []
            
           
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                name, enrollment,father_name, mother_name, mentoring_started_year, department, email, caste, weight, height, address, city, state, pincode, student_contact, father_contact, mother_contact, father_occupation, mother_occupation, hobbies, nationality, sport, overall_10th, overall_12th, diploma, study_semester,attendence, division,gender, mentorname, feespaid,fessunpaid,spi,cpi, *other_fields = row
                print("Name:", name)
                try:
                    
                    with transaction.atomic():
                        student = Student.objects.create(
                            name=name,
                            enrollment=enrollment,
                            # student_photo=student_photo,
                            study_semester=study_semester,
                            # father_photo=father_photo,
                            father_name=father_name,
                            mother_name=mother_name,
                            # mother_photo=mother_photo,
                            mentoring_started_year=mentoring_started_year,
                            department=Department.objects.filter(department_title=department).first(),
                            mentor_name=mentorname,
                            email=email,
                            caste=caste,
                            gender=gender,
                            weight=weight,
                            height=height,
                            address=address,
                            division=division,
                            city=city,
                            state=state,
                            pincode=pincode,
                            fees_paid=feespaid,
                            fees_unpaid=fessunpaid,
                            student_contact=student_contact,
                            father_contact=father_contact,
                            mother_contact=mother_contact,
                            father_occupation=father_occupation,
                            mother_occupation=mother_occupation,
                            nationality=nationality,
                            overall_10th=overall_10th,
                            overall_12th=overall_12th,
                            diploma=diploma,
                            attendence=attendence,
                            spi=spi,
                            cpi=cpi,
                            **{field_name: field_value for field_name, field_value in zip(other_fields)}
                        )

                        # Handle hobbies
                        hobby_objects = []
                        for hobby_name in hobbies.split(','):
                            hobby, created = Hobby.objects.get_or_create(title=hobby_name.strip())
                            hobby_objects.append(hobby)

                        student.hobbies.set(hobby_objects)
                        
                        
                        sport_objects = []
                        for sport_name in sport.split(','):
                            sport_obj, created = Sports.objects.get_or_create(sport_title=sport_name.strip())
                        sport_objects.append(sport_obj)

                        student.sport.set(sport_objects)

                        students_to_create.append(student)
                    # students_to_create.append(student)
                    # Student.objects.create(student)
                    

                except ValidationError as e:
                    # Log the validation error
                    print(f"Validation error for row {row}: {e}")
            Student.objects.bulk_create(students_to_create)
            wb.close()  # Close the workbook
            return HttpResponse(json.dumps({"msg": "Your details updated successfully."}), content_type="application/json")
        except Exception as e:
            # Log the exception
            print(f"An error occurred: {e}")
            return HttpResponse(json.dumps({"msg": "An error occurred while importing data."}), content_type="application/json")

    return HttpResponse(json.dumps({"msg": "Your details updated successfully."}), content_type="application/json")

@csrf_exempt
def signup(request):
    if request.method == "POST":
        # Fetch signup data
        username = request.POST.get('username')
        fname = request.POST.get('name')
        lname = request.POST.get('lname')
        email = request.POST.get('email')
        pass1 = request.POST.get('pass1')
        user_type = request.POST.get('user_type')
        print(user_type)

        # Validations (same as original, plus validate user_type)
        if User.objects.filter(username=username).exists():
            messages.error(request, "")
            return HttpResponse(json.dumps({"msg": "Username already exist! Please try some other username.","status":1}), content_type="application/json")
        
        if User.objects.filter(email=email).exists():
            messages.error(request, "Email Already Registered!!")
            return HttpResponse(json.dumps({"msg": "Email Already Registered!!","status":1}), content_type="application/json")

        if not user_type or user_type not in [choice[0] for choice in UserProfile.USER_CHOICES]:
            messages.error(request, "Please select a valid user type")
            return HttpResponse(json.dumps({"msg": "Please select a valid user type","status":1}), content_type="application/json")

        # Create basic Django User
        user = User.objects.create_user(username, email, pass1)
        user.first_name = fname
        user.last_name = lname
        print("save")
        user.save()

        usr2 = usr.User.objects.create_user(username, email, pass1)
        usr2.first_name = fname
        usr2.last_name = lname
        usr2.save()

        user_profile = UserProfile(user=user, user_type=user_type)
        user_profile.save()

        if user_type == 'Chairman':
            chairman = Chairman(user=user)
            chairman.name = username
            chairman.email = email
            chairman.save()

        elif user_type == 'Principal':
            principal = Principal(user=user)
            principal.name = username
            principal.email = email
            college = request.POST.get("college")
            colleges = College.objects.filter(name=college).first()
            principal.college = colleges
            principal.education = request.POST.get("education")
            principal.save()

        elif user_type == 'HOD':
            hod = HOD(user=user)
            print(user_type)
            hod.name = username
            hod.email = email
            department = request.POST.get("department")
            departments = Department.objects.filter(department_title=department).first()
            hod.department = departments
            hod.education = request.POST.get("education")
            hod.save()

        elif user_type == 'Mentor':
            mentor = Mentor(user=user)
            mentor.name = username
            mentor.email = email
            department = request.POST.get("department")
            departments = Department.objects.filter(department_title=department).first()
            mentor.department = departments
            mentor.mentoring_class = request.POST.get("mentoring_calss")
            mentor.number_of_students = request.POST.get("number_of_students")
            mentor.save()

        messages.success(request, "Your Account has been created succesfully!!")
        return HttpResponse(json.dumps({"msg": "Your details have been updated successfully.","status":1}), content_type="application/json")

    return render(json.dumps({"msg": "Your Account has been created succesfully!!","status":2}), content_type="application/json")



# to fetch students under specific mentor (mentor side)
@csrf_exempt
def students_mentor(request):
    if request.method == "POST":
        semester = request.POST.get("semester")
        mentor = request.POST.get("mentor")
        mentors = Mentor.objects.filter(name=mentor).first()
        # division = mentors.mentoring_class
        # division=division
        if mentors:
            studentsof = Student.objects.filter(study_semester=semester, mentor_name=mentors.name).all()
            data = serializers.serialize("json", studentsof)
            return HttpResponse(json.dumps({"data": data}), content_type="application/json")
        else:
            # Handle the case when mentor is not found
            return HttpResponse(json.dumps({"error": "Mentor not found"}), status=404)
    
# need to add a button and a file field in frontend and catch the response
def upload_file(request):
    if request.method == 'POST':
        feeeproof = Feesproof()
        student_name = request.POST.get("name")
        feeeproof.student = Student.objects.filter(name=student_name)
        feeeproof.proof = request.FILES.get("proof")
        feeeproof.save()

# to check if the stuent has paid fees or not in fees list in frontend
@csrf_exempt
def fess_check(request):
    if request.method == "POST":
        semester = request.POST.get("semester")  # Use GET for retrieving data
        students = Student.objects.filter(study_semester=semester).all()
        data = []
        for student in students:
            # Check if fees_paid is greater than 4000
            fees_unpaid = student.fees_unpaid  # Assuming there's a field named fees_paid in the Student model
            payment_status = "Paid in full" if fees_unpaid == 0 else "Not paid in full"
            
            # Serialize student data along with payment status
            student_data = serializers.serialize("json", [student])
            student_dict = {"student_data": student_data, "payment_status": payment_status}
            data.append(student_dict)
        
        return HttpResponse(json.dumps({"data": data}),content_type="application/json")
    
#for hod side student list
@csrf_exempt
def students(request):
    if request.method == "POST":
        semester = request.POST.get("semester")
        studentsof = Student.objects.filter(study_semester=1).all()
        data = serializers.serialize("json", studentsof)
        return HttpResponse(json.dumps({"data": data}),content_type="application/json")
    else:
        HttpResponse(json.dumps({"data": "Invalid req"}),content_type="application/json")


@csrf_exempt
def signin(request):
      if request.method == 'POST':
        username = request.POST['username']
        pass1 = request.POST['pass1']
        
        user = authenticate(username=username, password=pass1)
        
        if user is not None:
            login(request, user)
            fname = user.first_name
            return HttpResponse(json.dumps({"msg": " Sucessfully Logged In ","status":True, "user_type" : user.user_type}),content_type="application/json",)
            messages.success(request, "Logged In Sucessfully!!")
            
        else:
            messages.error(request, "Bad Credentials!!")
    
      return HttpResponse(json.dumps({"msg": " Bad Request ","status":False}),content_type="application/json",)


def colleges(request):
    if request.method == "GET":
        colleges = College.objects.all()
        data = serializers.serialize("json", colleges)
        return HttpResponse(json.dumps({"data": data}), content_type="application/json")
    
    
@csrf_exempt
def mentors(request):
    if request.method == "POST":
        college = request.POST.get("college")
        department = request.POST.get("department")
        colleges = College.objects.filter(name=college).first()
        departments = Department.objects.filter(department_title=department).first()
        mentors = Mentor.objects.filter(college=colleges.pk, department=departments.pk)
        data = serializers.serialize("json", mentors)
        return HttpResponse(json.dumps({"data": data}), content_type="application/json")
    
@csrf_exempt
def departments(request):
    if request.method == "POST":
        college = request.POST.get("college")
        colleges = College.objects.filter(name=college).first()
        departments = Department.objects.filter(college=colleges.pk).all()
        data = serializers.serialize("json", departments)
        return HttpResponse(json.dumps({"data": data}), content_type="application/json")
    
    

    
    
    
@csrf_exempt 
def result(request):
 if request.method == "POST":
        semester = request.POST.get("semester")
        # type_of_exam = request.POST.get("type_of_exam")
        student = request.POST.get("student")
#  type_of_exam=type_of_exam, 
        students = Student.objects.filter(name=student).first()
        results = Result.objects.filter(semester=semester,student=students.pk).all()
        data = []
        for result in results:
            result_data = {
                "semester": result.semester,
                "type_of_exam": result.type_of_exam,
                "student": result.student.name,
                "subject": result.subject.subject_name,
                "currriculum_year": result.currriculum_year,
                "marks": result.marks,
                "grade": result.grade,
                "exam_year": result.exam_year
            }
            data.append(result_data)
        return HttpResponse(json.dumps({"data" : data}), content_type="application/json")


@csrf_exempt 
def addStudent(request):
    if request.method == "POST":
        studet_instance = Student()
        studet_instance.name = request.POST.get('name')
        studet_instance.enrollment = request.POST.get('enrollment')
       
        studet_instance.student_photo = request.FILES.get('student_photo')
        studet_instance.father_photo = request.FILES.get('father_photo')
        studet_instance.mother_photo = request.FILES.get('mother_photo')
        studet_instance.mentoring_started_year = request.POST.get('mentoring_started_year')
        departments = request.POST.get('departments')
        department = Department.objects.filter(department_title=departments).first()
        studet_instance.department = department
        studet_instance.division = request.POST.get('division')
        studet_instance.date_of_birth = request.POST.get('date_of_birth')
        studet_instance.height = request.POST.get('height')
        studet_instance.weight = request.POST.get('weight')
        studet_instance.gender = request.POST.get('gender')
        studet_instance.caste = request.POST.get('caste')
        studet_instance.address = request.POST.get('address')
        studet_instance.city = request.POST.get('city')
        studet_instance.state = request.POST.get('state')
        studet_instance.nationality = request.POST.get('country')
        studet_instance.overall_10th = request.POST.get('overall_10th')
        studet_instance.overall_12th = request.POST.get('overall_12th')
        studet_instance.diploma = request.POST.get('diploma')
        studet_instance.student_contact = request.POST.get('student_contact')
        studet_instance.father_contact = request.POST.get('father_contact')
        studet_instance.mother_contact = request.POST.get('mother_contact')
        studet_instance.father_occupation = request.POST.get('father_occupation')
        studet_instance.mother_occupation = request.POST.get('mother_occupation')
        studet_instance.father_name = request.POST.get('father_name')
        studet_instance.mother_name = request.POST.get('mother_name')
        studet_instance.activity = request.POST.get('activity')
        studet_instance.project = request.POST.get('project')
        studet_instance.membership_in_organization = request.POST.get('membership_in_organization')
        studet_instance.attendence = request.POST.get('attendence')
        studet_instance.pincode = request.POST.get('pincode')
        studet_instance.save()

        return HttpResponse(json.dumps({"msg": "Success"}))
    else:
        return HttpResponse(json.dumps({"msg": "Bad request"}))
    
@csrf_exempt
def resultExcel(request):
    if request.method == "POST":
        file = request.FILES.get('excel_file')
        if not file:
            return HttpResponse(json.dumps({"Error": "No file uploaded"}), status=400)
        try:
            df = pd.read_excel(file)
            subject_columns = [col for col in df.columns if col.startswith('subject')]
            for index, row in df.iterrows():
                student_name = row['student']
                exam_type = row['exam type']
                semester = row['semester']
                currriculum_year = row['currriculum_year']
                exam_year = row['exam_year']
                grade = row['grade']
                marks = row['marks']
                
                # Handling multiple students with the same name
                students = Student.objects.filter(name=student_name)
                if students.exists():
                    student = students.first()  # Get the first student with the given name
                else:
                    student = Student.objects.create(name=student_name)
                for subject_column in subject_columns:
                    subject_name = row[subject_column]
                    subject_name = subject_column.replace("subject ","")
                  
                  
                    try:  
                        subject_obj = Subject.objects.filter(subject_name=subject_name).first()
                        result, created = Result.objects.get_or_create(student=student, subject=subject_obj, type_of_exam=exam_type, semester=semester,grade=grade, marks=marks ,exam_year=exam_year, currriculum_year=currriculum_year)
                        print(subject_columns)
                        # ['subject Dhiraj Suthar']
                        print(f"Processed {student_name}, Subject: {subject_name}, Exam Type: {exam_type}")
                    except Result.MultipleObjectsReturned:
                        print(subject_columns)
                        subject_obj = Subject.objects.filter(subject_name=subject_name).first()
                        Result.objects.filter(student=student, subject=subject_obj, type_of_exam=exam_type).delete()
                        result, created = Result.objects.get_or_create(student=student, subject=subject_obj, grade=grade, marks=marks  , type_of_exam=exam_type, semester=semester, exam_year=exam_year, currriculum_year=currriculum_year)
                        print(f"Processed {student_name}, Subject: {subject_name}, Exam Type: {exam_type}")
            return HttpResponse(json.dumps({"Message": "Success"}))
        except Exception as e:
            print(f"Error: {e}")
            return HttpResponse(json.dumps({"Error": str(e)}), status=500)
    else:
        return HttpResponse(json.dumps({"Error": "Only POST requests are allowed"}), status=405)

@csrf_exempt
def createStudentProfile(request):
    if request.method == "POST":
        enrollment = request.POST.get("enrollment")
        password = request.POST.get("password")

        users = Student.objects.get(enrollment=enrollment)

        studentProfile = StudentProfile.objects.create(
            username=enrollment,
            student_instance=users,
            password=password
        )

        return HttpResponse(json.dumps({"Message": "Student Profile Created"}))
    else:
        return HttpResponse(json.dumps({"Message": "Student Profile Not Created"}))
    
@csrf_exempt
def studentLogin(request):
    if request.method == "POST":
        enrollment = request.POST.get("enrollment")
        password = request.POST.get("password")

        user = StudentProfile.objects.filter(username=enrollment, password=password).exists()

        if user:
            return HttpResponse(json.dumps({"Message" : "Logged In"}))
        else:
            return HttpResponse(json.dumps({"Message" : "Invalid Creddentials"}))
    else:
        return HttpResponse(json.dumps({"Message" : "Bad Request"}))
    
@csrf_exempt
def editStudent(request):
    if request.method == "POST":
        enrollment = request.POST.get("enrollment")
        try:
            student_instance = Student.objects.get(enrollment=enrollment)
        except Student.DoesNotExist:
            return JsonResponse({"msg": "Student does not exist"}, status=404)
        data = json.loads(request.POST.get("data"))
        for field, value in data.items():
            setattr(student_instance, field, value)
        student_instance.save()
        return HttpResponse(json.dumps({"msg": "Success"}))
    else:
        return HttpResponse(json.dumps({"msg": "Bad request"}))

@csrf_exempt
def passOutTick(request):
    students = Student.objects.all()
    current_year = datetime.datetime.now().year
    
    for student in students:
        nums = student.department.number_of_semesters / 2
        student.mentoring_started_year.year + nums >= current_year
        student.passed_out = True
        student.save()
    return HttpResponse(json.dumps({'success': 'True'}))
    
